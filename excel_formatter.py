import os
from typing import List, Optional
import pandas as pd
import xlsxwriter
from xlsxwriter.workbook import Workbook
from xlsxwriter.worksheet import Worksheet
import openpyxl
from datetime import datetime
from config import Config
from data_processor import SalesData


class ExcelFormatter:
    """Handles Excel file creation and formatting"""

    def __init__(self, config: Config):
        self.config = config
        self.filedate = datetime.now().strftime("%y%m%d-%H%M%S")

    def create_reports(self, sales_data: SalesData) -> dict[str, str]:
        """
        Create Excel reports for each AE

        Args:
            sales_data: Processed sales data

        Returns:
            Dictionary mapping AE names to their report file paths
        """
        reports_created = {}

        for ae_name in sales_data.report.AE1.unique():
            report_path = self._create_single_report(ae_name, sales_data)
            reports_created[ae_name] = report_path

        return reports_created

    def _create_single_report(self, ae_name: str, sales_data: SalesData) -> str:
        """Create a single AE's report"""
        # Generate filename and path
        filename = f"{ae_name}-Sales Tool-{self.filedate}.xlsm"
        full_path = os.path.join(self.config.reports_folder, filename)

        try:
            # First, create with openpyxl for data
            self._create_initial_workbook(full_path, ae_name, sales_data)

            # Then, reopen with xlsxwriter for formatting and VBA
            self._format_workbook(full_path, ae_name, sales_data)

            return full_path

        except Exception as e:
            raise RuntimeError(f"Error creating report for {ae_name}: {str(e)}") from e

    def _create_initial_workbook(
        self, filepath: str, ae_name: str, sales_data: SalesData
    ):
        """Create initial workbook with data"""
        with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
            # Filter and write main report
            temp_df = sales_data.report[sales_data.report.AE1 == ae_name]
            temp_df.to_excel(writer, sheet_name="Sheet1", index=False)

            # Filter and write budget report
            budget_df = sales_data.budget_unassigned[
                sales_data.budget_unassigned["AE1"] == ae_name
            ]
            budget_df.to_excel(
                writer, sheet_name="Budget-Assigned-Unassigned", index=False
            )

    def _format_workbook(self, filepath: str, ae_name: str, sales_data: SalesData):
        """Add formatting and VBA to workbook"""
        # Create new workbook with xlsxwriter
        workbook = xlsxwriter.Workbook(filepath, {"vba_codename": "ThisWorkbook"})

        try:
            # Read data from saved file
            xlsx_file = openpyxl.load_workbook(filepath)

            # Format each sheet
            for sheet_name in xlsx_file.sheetnames:
                worksheet = workbook.add_worksheet(sheet_name)

                if sheet_name == "Sheet1":
                    worksheet.set_vba_name("Sheet1")

                # Copy data from saved file
                xlsx_sheet = xlsx_file[sheet_name]
                for row_idx, row in enumerate(xlsx_sheet.iter_rows(values_only=True)):
                    worksheet.write_row(row_idx, 0, row)

                # Apply appropriate formatting
                if sheet_name == "Sheet1":
                    self._format_sheet1(workbook, worksheet, sales_data)
                else:
                    self._format_sheet2(workbook, worksheet, sales_data)

            # Add VBA project
            workbook.add_vba_project(self.config.vba_path)

        finally:
            workbook.close()

    def _format_sheet1(self, workbook: Workbook, worksheet: Worksheet, sales_data: SalesData):
        """Format the main sales report sheet"""
        money_fmt = workbook.add_format({"num_format": 42, "align": "center"})
        text_fmt = workbook.add_format({"align": "left"})

        worksheet.set_column("A:B", 15, text_fmt)
        worksheet.set_column("C:C", 30, text_fmt)
        worksheet.set_column("D:G", 10, money_fmt)

        # Filter data for current AE
        ae_data = sales_data.report[sales_data.report['AE1'] == ae_name]
        
        # Calculate table range including all rows plus header and total row
        total_rows = len(ae_data)
        table_range = f"A1:G{total_rows + 1}"  # +1 for header, +1 for total row

        columns = [{"header": "AE1"}, {"header": "Sector"}, {"header": "Customer"}]
        for quarter in sales_data.quarter_columns:
            columns.append({"header": quarter, "total_function": "sum"})

        worksheet.add_table(
            table_range,
            {
                "columns": columns,
                "autofilter": True,
                "total_row": True,
                "style": "Table Style Light 11",
            }
        )

        worksheet.freeze_panes(1, 0)
        worksheet.set_zoom(90)

    def _format_sheet2(
        self, workbook: Workbook, worksheet: Worksheet, sales_data: SalesData
    ):
        """Format the budget and unassigned sheet"""
        money_fmt = workbook.add_format({"num_format": 42, "align": "center"})
        text_fmt = workbook.add_format({"align": "left"})

        # Set column formats
        worksheet.set_column("A:B", 15, text_fmt)
        worksheet.set_column("C:C", 30, text_fmt)
        worksheet.set_column("D:H", 10, money_fmt)

        # Add formatting options
        worksheet.freeze_panes(1, 0)
        worksheet.set_zoom(90)

    def _add_sales_table(self, worksheet: Worksheet, sales_data: SalesData):
        """Add formatted table to sales sheet"""
        # Calculate table dimensions
        row_length = len(sales_data.report.index) + 2
        table_range = f"A1:G{row_length}"

        # Create column definitions
        columns = [{"header": "AE1"}, {"header": "Sector"}, {"header": "Customer"}]

        # Add quarter columns
        for quarter in sales_data.quarter_columns:
            columns.append({"header": quarter, "total_function": "sum"})

        # Add table
        worksheet.add_table(
            table_range,
            {
                "columns": columns,
                "autofilter": True,
                "total_row": True,
                "style": "Table Style Light 11",
            },
        )

    def _add_vba_button(self, worksheet: Worksheet):
        """Add VBA macro button to worksheet"""
        worksheet.insert_button(
            "J8",
            {
                "macro": "Sheet1.CreateSlicerVBA",
                "caption": "Press Me",
                "width": 80,
                "height": 30,
            },
        )
